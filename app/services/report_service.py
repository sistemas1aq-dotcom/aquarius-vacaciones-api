"""Reporte Corporativo de Vacaciones: la réplica del formato en Excel.

Las tres salidas — vista previa en pantalla, XLSX y PDF — se arman a partir de
UNA sola función, `datos_reporte()`. Es deliberado: si cada formato calculara
lo suyo, tarde o temprano el Excel diría una cosa y la pantalla otra, y el que
se manda a gerencia es el Excel.

SOBRE EL AÑO Y LO QUE SIGNIFICA CADA COLUMNA
--------------------------------------------
El selector de año NO recalcula el saldo a esa fecha. Cambia únicamente el
reparto de los días gozados entre las dos columnas:

    Vacaciones Ganadas x Año          -> EarnedDays del saldo (a HOY)
    Días de Vac. Gozadas - 31/12/AA   -> gozados ANTES del 1 de enero del año
    Días de Vac. Gozadas - AAAA       -> gozados DENTRO del año elegido
    Vac. Pendientes x Año             -> PendingByYear del saldo (a HOY)
    Vac. Truncos (Periodo actual)     -> PendingTruncated del saldo (a HOY)

Las columnas de saldo salen del último cálculo de `VacationBalances`, el mismo
que alimenta el dashboard y los recordatorios. Se hace así a propósito: un
reporte que contradiga a la pantalla de al lado no sirve para nada, aunque sus
cifras estén mejor razonadas. Reconstruir el saldo histórico a 31/12 de un año
pasado es otro trabajo, y habría que decidir antes qué se hace con las
vacaciones registradas después de esa fecha.

Los gozados sí se calculan contra la tabla `Vacations` con el año pedido, no
con `TakenDays2026`, que está fijado a ese año en el procedimiento almacenado.
"""

from __future__ import annotations

import io
from datetime import date

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.models import Department, Employee, Vacation, VacationBalance

#: Estados que cuentan como días gozados. Los mismos que usa
#: `sp_calculate_vacation_balance`: si aquí se contara distinto, el reporte
#: discreparía del saldo que muestra la aplicación.
ESTADOS_GOZADOS = ("completed", "in_progress", "approved")


def _num(valor) -> float:
    return float(valor) if valor is not None else 0.0


def datos_reporte(db: Session, year: int) -> dict:
    """Filas y totales del reporte, en el orden del formato corporativo."""
    # Último saldo de cada empleado. Se resuelve con una subconsulta de máximos
    # en vez de una consulta por empleado: con 100 trabajadores la diferencia
    # es irrelevante, pero el patrón "una consulta por fila" es el que acaba
    # doliendo cuando la plantilla crece.
    ultimo = (
        db.query(
            VacationBalance.EmployeeId.label("emp"),
            func.max(VacationBalance.CalculationDate).label("fecha"),
        )
        .group_by(VacationBalance.EmployeeId)
        .subquery()
    )

    filas_db = (
        db.query(Employee, Department, VacationBalance)
        .join(Department, Employee.DepartmentId == Department.Id)
        .outerjoin(ultimo, ultimo.c.emp == Employee.Id)
        .outerjoin(
            VacationBalance,
            (VacationBalance.EmployeeId == Employee.Id)
            & (VacationBalance.CalculationDate == ultimo.c.fecha),
        )
        .filter(Employee.IsActive == True)  # noqa: E712 - SQLAlchemy
        .order_by(Department.Name, Employee.FullName)
        .all()
    )

    inicio = date(year, 1, 1)
    fin = date(year, 12, 31)

    # Gozados por empleado, en dos tramos, de una sola pasada.
    gozados_ano = dict(
        db.query(Vacation.EmployeeId, func.sum(Vacation.Days))
        .filter(Vacation.Status.in_(ESTADOS_GOZADOS),
                Vacation.StartDate >= inicio,
                Vacation.StartDate <= fin)
        .group_by(Vacation.EmployeeId)
        .all()
    )
    gozados_antes = dict(
        db.query(Vacation.EmployeeId, func.sum(Vacation.Days))
        .filter(Vacation.Status.in_(ESTADOS_GOZADOS),
                Vacation.StartDate < inicio)
        .group_by(Vacation.EmployeeId)
        .all()
    )

    filas = []
    for i, (emp, depto, saldo) in enumerate(filas_db, start=1):
        filas.append({
            "N": i,
            "Dni": emp.Dni or "",
            "FullName": emp.FullName,
            "Department": depto.Name if depto else "",
            # El formato corporativo trae una columna DETALLE que hoy no tiene
            # origen en la base. Se emite vacía en vez de inventarle contenido.
            "Detalle": "",
            "Position": emp.Position or "",
            "HireDate": emp.HireDate.isoformat() if emp.HireDate else None,
            "Earned": _num(saldo.EarnedDays) if saldo else 0.0,
            "TakenPrev": _num(gozados_antes.get(emp.Id)),
            "TakenYear": _num(gozados_ano.get(emp.Id)),
            "PendingByYear": _num(saldo.PendingByYear) if saldo else 0.0,
            "Truncated": _num(saldo.PendingTruncated) if saldo else 0.0,
        })

    def suma(campo: str) -> float:
        return round(sum(f[campo] for f in filas), 2)

    return {
        "year": year,
        "previousYearShort": f"{(year - 1) % 100:02d}",
        "rows": filas,
        "totals": {
            "Earned": suma("Earned"),
            "TakenPrev": suma("TakenPrev"),
            "TakenYear": suma("TakenYear"),
            "PendingByYear": suma("PendingByYear"),
            "Truncated": suma("Truncated"),
        },
    }


# ───────────────────────────────────────────────────────────────────
# Cabeceras: una sola definición para las tres salidas
# ───────────────────────────────────────────────────────────────────

def _cabeceras(datos: dict) -> list[tuple[str, str]]:
    """(texto, grupo de color) en el orden del formato corporativo."""
    año, prev = datos["year"], datos["previousYearShort"]
    return [
        ("Nº", "azul"),
        ("DNI", "azul"),
        ("APELLIDOS Y NOMBRES", "azul"),
        ("PLANILLA", "azul"),
        ("DETALLE", "azul"),
        ("CARGO", "azul"),
        ("FECHA DE INGRESO", "azul"),
        ("Vacaciones Ganadas x Año", "gris"),
        (f"Dias de Vac. Gozadas - 31/12/{prev}", "gris"),
        (f"Dias de Vac. Gozadas - {año}", "gris"),
        ("Vac. Pendientes x Año", "celeste"),
        ("Vac. Truncos (Periodo actual)", "celeste"),
    ]


def _valores(fila: dict) -> list:
    fecha = fila["HireDate"]
    if fecha:
        d = date.fromisoformat(fecha)
        fecha = f"{d.day:02d}/{d.month:02d}/{d.year}"
    return [
        fila["N"], fila["Dni"], fila["FullName"], fila["Department"],
        fila["Detalle"], fila["Position"], fecha or "",
        fila["Earned"], fila["TakenPrev"], fila["TakenYear"],
        fila["PendingByYear"], fila["Truncated"],
    ]


# ───────────────────────────────────────────────────────────────────
# XLSX
# ───────────────────────────────────────────────────────────────────

def generar_xlsx(datos: dict) -> bytes:
    """Libro de Excel con los colores del formato corporativo."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    RELLENOS = {
        "azul":    PatternFill("solid", fgColor="0000CC"),
        "gris":    PatternFill("solid", fgColor="3B618E"),
        "celeste": PatternFill("solid", fgColor="B7DEE8"),
    }
    LETRA = {
        "azul":    Font(bold=True, color="FFFFFF", size=9),
        "gris":    Font(bold=True, color="FFFFFF", size=9),
        "celeste": Font(bold=True, color="FF0000", size=9),
    }
    TOTAL = PatternFill("solid", fgColor="FFF2CC")
    linea = Side(style="thin", color="000000")
    borde = Border(left=linea, right=linea, top=linea, bottom=linea)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Vacaciones {datos['year']}"

    cabeceras = _cabeceras(datos)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabeceras))
    titulo = ws.cell(row=1, column=1, value=f"REPORTE DE VACACIONES — {datos['year']}")
    titulo.font = Font(bold=True, size=12)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Cabecera
    for col, (texto, grupo) in enumerate(cabeceras, start=1):
        c = ws.cell(row=2, column=col, value=texto)
        c.fill = RELLENOS[grupo]
        c.font = LETRA[grupo]
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borde
    ws.row_dimensions[2].height = 34

    # Datos
    for i, fila in enumerate(datos["rows"], start=3):
        for col, valor in enumerate(_valores(fila), start=1):
            c = ws.cell(row=i, column=col, value=valor)
            c.border = borde
            c.font = Font(size=9)
            if col in (1, 2, 4, 5, 7, 8, 9, 10, 11, 12):
                c.alignment = Alignment(horizontal="center")
            if col == 12:
                c.number_format = "0.00"

    # Totales
    ultima = len(datos["rows"]) + 3
    if datos["rows"]:
        ws.cell(row=ultima, column=7, value="TOTAL").alignment = Alignment(horizontal="right")
        totales = [
            datos["totals"]["Earned"], datos["totals"]["TakenPrev"],
            datos["totals"]["TakenYear"], datos["totals"]["PendingByYear"],
            datos["totals"]["Truncated"],
        ]
        for col, valor in enumerate(totales, start=8):
            ws.cell(row=ultima, column=col, value=valor)
        for col in range(1, len(cabeceras) + 1):
            c = ws.cell(row=ultima, column=col)
            c.fill = TOTAL
            c.font = Font(bold=True, size=9)
            c.border = borde
            if col >= 8:
                c.alignment = Alignment(horizontal="center")
        ws.cell(row=ultima, column=12).number_format = "0.00"

    anchos = [5, 12, 34, 18, 12, 30, 14, 12, 14, 14, 12, 14]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Fija título y cabecera al desplazarse: un reporte de 100 filas sin esto
    # se recorre sin saber qué columna se está mirando.
    ws.freeze_panes = "A3"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ───────────────────────────────────────────────────────────────────
# PDF
# ───────────────────────────────────────────────────────────────────

def generar_pdf(datos: dict) -> bytes:
    """PDF apaisado. Doce columnas no caben en vertical."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=f"Reporte de Vacaciones {datos['year']}",
    )

    estilos = getSampleStyleSheet()
    normal = estilos["BodyText"]
    normal.fontSize = 6.5
    normal.leading = 8

    cabeceras = _cabeceras(datos)
    tabla = [[Paragraph(f"<b>{t}</b>", normal) for t, _ in cabeceras]]
    for fila in datos["rows"]:
        valores = _valores(fila)
        tabla.append([
            Paragraph(str(v if v is not None else ""), normal) for v in valores
        ])

    if datos["rows"]:
        t = datos["totals"]
        tabla.append([
            Paragraph("", normal)] * 6 + [Paragraph("<b>TOTAL</b>", normal)] + [
            Paragraph(f"<b>{t['Earned']:.0f}</b>", normal),
            Paragraph(f"<b>{t['TakenPrev']:.0f}</b>", normal),
            Paragraph(f"<b>{t['TakenYear']:.0f}</b>", normal),
            Paragraph(f"<b>{t['PendingByYear']:.0f}</b>", normal),
            Paragraph(f"<b>{t['Truncated']:.2f}</b>", normal),
        ])

    # Anchos RELATIVOS, repartidos sobre el ancho útil de la página. Fijarlos
    # en milímetros absolutos es lo que hacía que la suma se pasara del papel
    # y reportlab reventara con un ancho negativo.
    proporciones = [12, 40, 105, 55, 35, 90, 42, 36, 42, 42, 36, 42]
    total = sum(proporciones)
    anchos = [doc.width * p / total for p in proporciones]
    tabla_pdf = Table(tabla, colWidths=anchos, repeatRows=1)

    estilo = [
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#666666")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Con doce columnas el relleno por defecto (6 pt por lado) se come
        # las estrechas enteras.
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("BACKGROUND", (0, 0), (6, 0), colors.HexColor("#0000CC")),
        ("BACKGROUND", (7, 0), (9, 0), colors.HexColor("#3B618E")),
        ("BACKGROUND", (10, 0), (11, 0), colors.HexColor("#B7DEE8")),
        ("TEXTCOLOR", (0, 0), (9, 0), colors.white),
        ("TEXTCOLOR", (10, 0), (11, 0), colors.red),
        ("ALIGN", (7, 1), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (1, -1), "CENTER"),
        ("ALIGN", (4, 1), (4, -1), "CENTER"),
        ("ALIGN", (6, 1), (6, -1), "CENTER"),
    ]
    if datos["rows"]:
        estilo.append(("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF2CC")))
    tabla_pdf.setStyle(TableStyle(estilo))

    titulo = estilos["Title"]
    titulo.fontSize = 13
    doc.build([
        Paragraph(f"REPORTE DE VACACIONES — {datos['year']}", titulo),
        Spacer(1, 4 * mm),
        tabla_pdf,
    ])
    return buffer.getvalue()
